from __future__ import annotations

import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from ..models import ColumnDefinition
from ..variables.var import EXPORT_SHEET_NAME


HEADER_FILL = PatternFill(fill_type="solid", start_color="1E2235", end_color="1E2235")
HEADER_FONT = Font(bold=True, color="FFFFFF")
PRESENCE_FILL = PatternFill(fill_type="solid", start_color="EF4444", end_color="EF4444")
ABSENCE_FILL = PatternFill(fill_type="solid", start_color="22C55E", end_color="22C55E")
UNKNOWN_FILL = PatternFill(fill_type="solid", start_color="6B7280", end_color="6B7280")
STATUS_FONT = Font(bold=True, color="FFFFFF")

NO_COLUMN_KEY = "no"
NBR_PRELEVEMENTS_KEY = "nbr_prelevements"
RESERVES_KEY = "reserves"
MATERIAUX_AMIANTES_KEY = "materiaux_amiantes"
LOCALISATION_KEY = "localisation"
PRESENCE_COLUMN_KEYS = {
    "cuisine_sol",
    "cuisine_murs",
    "cuisine_plafond",
    "cuisine_faience",
    "cuisine_evier",
    "salle_bain_sol",
    "salle_bain_murs",
    "salle_bain_plafonds",
    "salle_bain_faience",
    "wc_sol",
    "wc_murs",
    "wc_plafond",
    "loggia_balcon",
    "celliers",
    "autre",
}
UNKNOWN_VALUE_MARKERS = {
    "",
    "?",
    "n/a",
    "na",
    "nd",
    "non precise",
    "non precis",
    "non renseigne",
    "non renseignee",
    "non renseigné",
    "non renseignée",
    "non spécifié",
    "non specifie",
    "non spécifiée",
    "non specifiee",
}


def strip_accents(value: str) -> str:
    return "".join(
        character
        for character in unicodedata.normalize("NFD", value)
        if unicodedata.category(character) != "Mn"
    )


def normalize_free_text(value: object) -> str | None:
    if value is None:
        return None

    normalized_value = str(value).strip()
    return normalized_value or None


def normalize_presence_value(value: object) -> str:
    text_value = normalize_free_text(value)

    if text_value is None:
        return "?"

    normalized_marker = strip_accents(text_value).lower().replace("’", "'")

    if normalized_marker in UNKNOWN_VALUE_MARKERS:
        return "?"

    if "absence" in normalized_marker:
        return "Absence"

    if "presence" in normalized_marker:
        return "Présence"

    return "?"


def normalize_nbr_prelevements(value: object) -> str:
    text_value = normalize_free_text(value)

    if text_value is None:
        return "ND"

    digits = "".join(character for character in text_value if character.isdigit())
    return digits or "ND"


def normalize_column_value(column: ColumnDefinition, value: object, row_number: int | None = None) -> str | None:
    if column.key == NO_COLUMN_KEY:
        if row_number is not None:
            return str(row_number)

        return normalize_free_text(value)

    if column.key == NBR_PRELEVEMENTS_KEY:
        return normalize_nbr_prelevements(value)

    if column.key == RESERVES_KEY:
        return normalize_free_text(value) or "Aucune réserve mentionnée"

    if column.key == MATERIAUX_AMIANTES_KEY:
        return normalize_free_text(value) or "Aucun matériau identifié"

    if column.key == LOCALISATION_KEY:
        return normalize_free_text(value) or "Aucune localisation identifiée"

    if column.key in PRESENCE_COLUMN_KEYS:
        return normalize_presence_value(value)

    return normalize_free_text(value)


def normalize_extracted_row(
    row: dict[str, str | None],
    columns: list[ColumnDefinition],
    *,
    row_number: int | None = None,
) -> dict[str, str | None]:
    normalized_row: dict[str, str | None] = {}

    for column in columns:
        normalized_row[column.key] = normalize_column_value(
            column,
            row.get(column.key),
            row_number=row_number,
        )

    return normalized_row


def normalize_dataframe_for_export(
    dataframe: pd.DataFrame,
    columns: list[ColumnDefinition],
) -> pd.DataFrame:
    normalized_dataframe = dataframe.copy()

    for row_index in range(len(normalized_dataframe.index)):
        for column in columns:
            if column.label not in normalized_dataframe.columns:
                continue

            normalized_dataframe.at[row_index, column.label] = normalize_column_value(
                column,
                normalized_dataframe.at[row_index, column.label],
                row_number=row_index + 1,
            )

    return normalized_dataframe


def build_dataframe(
    rows: list[dict[str, str | None]],
    columns: list[ColumnDefinition],
) -> pd.DataFrame:
    ordered_labels = [column.label for column in columns]
    normalized_rows: list[dict[str, str | None]] = []

    for row_index, row in enumerate(rows, start=1):
        normalized_row = normalize_extracted_row(row, columns, row_number=row_index)
        normalized_rows.append({column.label: normalized_row.get(column.key) for column in columns})

    dataframe = pd.DataFrame(normalized_rows)

    for label in ordered_labels:
        if label not in dataframe.columns:
            dataframe[label] = None

    return normalize_dataframe_for_export(dataframe[ordered_labels], columns)


def read_export_dataframe(output_path: str) -> pd.DataFrame:
    output_file = Path(output_path)

    if not output_file.exists():
        raise FileNotFoundError(f"Fichier Excel introuvable: {output_file}")

    dataframe = pd.read_excel(
        output_file,
        sheet_name=EXPORT_SHEET_NAME,
        dtype=object,
    )

    return dataframe.astype(object)


def export_excel(
    dataframe: pd.DataFrame,
    output_path: str,
    columns: list[ColumnDefinition] | None = None,
) -> None:
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    export_dataframe = normalize_dataframe_for_export(dataframe, columns or [])

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        export_dataframe.to_excel(writer, index=False, sheet_name=EXPORT_SHEET_NAME)
        worksheet = writer.sheets[EXPORT_SHEET_NAME]
        worksheet.auto_filter.ref = worksheet.dimensions
        header_by_label = {column.label: column for column in (columns or [])}

        for header_cell in worksheet[1]:
            header_cell.fill = HEADER_FILL
            header_cell.font = HEADER_FONT
            header_cell.alignment = Alignment(vertical="center")

        worksheet.row_dimensions[1].height = 30

        for row_index in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row_index].height = 20

        if columns:
            for column_index, label in enumerate(export_dataframe.columns, start=1):
                column_definition = header_by_label.get(label)

                if column_definition is None or column_definition.key not in PRESENCE_COLUMN_KEYS:
                    continue

                for row_index in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_index, column=column_index)
                    normalized_value = normalize_presence_value(cell.value)
                    cell.value = normalized_value
                    cell.font = STATUS_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    if normalized_value == "Présence":
                        cell.fill = PRESENCE_FILL
                    elif normalized_value == "Absence":
                        cell.fill = ABSENCE_FILL
                    else:
                        cell.fill = UNKNOWN_FILL

        for column in worksheet.columns:
            max_length = max(
                (len(str(cell.value)) if cell.value is not None else 0 for cell in column),
                default=10,
            )
            worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 4, 60)
